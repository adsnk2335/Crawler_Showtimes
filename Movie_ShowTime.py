import random
import requests as req
from bs4 import BeautifulSoup
import re
import os
import openpyxl as xl
import time

path = r"C:\Users\ADSNK2335\Desktop\movie_showtime\\"
filename = "movie_showtime.xlsx"
if not os.path.exists(path):
    os.makedirs(path)
if os.path.exists(path+filename):
    os.remove(path+filename)
font_header = xl.styles.Font(name='微軟正黑體', size=14, bold=True)
font_other = xl.styles.Font(name='微軟正黑體', size=12)


def minshan_movie_info():
    home_page = "http://www.minshen.com.tw/"
    url = "http://www.minshen.com.tw/movies.php"
    headers = {"User-Agent": "Mozilla/5.0 ......."}
    resp = req.get(url, headers=headers)
    if resp.status_code == 200:
        resp.encoding = "UTF-8"
        soup = BeautifulSoup(resp.text, "html.parser")
        # print(soup.prettify()) # 輸出排版後的HTML程式碼
        movie_info = soup.findAll("div", class_="movie-info")
        cnt = 0
        t = 0
        rest = range(1, 6)
        for info in movie_info:
            new_url = home_page+info.div.a["href"]
            new_resp = req.get(new_url, headers=headers)
            if new_resp.status_code == 200:
                new_resp.encoding = "UTF-8"
                new_soup = BeautifulSoup(new_resp.text, "html.parser")
                title = new_soup.find("p", class_="chinese-title")  # 電影名稱
                show_info = new_soup.find("div", class_="playdate")
                show_date_split = re.split("\|", show_info.text)
                show_date = show_date_split[0][5:15].strip()    # 上映日期
                duration = show_date_split[1][4:].strip()   # 片長
                style = new_soup.find("div", class_="movie-more-information")
                type_td = style.table.tr.findAll("td")
                typee = type_td[1].text # 類型
                movie_desc = new_soup.find("div", class_="movie-description")
                desc = movie_desc.findAll("p")
                description = desc[1].text  # 劇情介紹
                trai_url = new_soup.find("div", class_="movie-trailer")
                trailer = trai_url.iframe["src"]    # 預告片
                movie_info_list = [title.text, show_date, duration, typee, description, trailer,new_url]
                detail_t = new_soup.find("div", class_="showtimes")
                detail_dt = detail_t.table.tbody.findAll("tr")
                detail_info(title.text, detail_dt, movie_info_list)
                cnt += 1
                time.sleep(random.choice(rest))
                exec_t = time.perf_counter()
                print(f"休息一下，我不是機器人，下載第{cnt}個資訊中，執行了{exec_t - t:.2f}秒")
                t = exec_t
            else:
                print("找不到該網頁")
        ad_font()
    else:
        print("找不到時刻表網頁")

def detail_info(title, tr, movie_info_list):
    if os.path.exists(path+filename):
        wb = xl.load_workbook(path+filename)
        ws = wb.active
    else:
        wb = xl.Workbook()
        ws = wb.active
        ws.append(["電影名稱", "上映日期", "片長", "類型", "劇情介紹", "電影預告", "戲院連結"])
    ws.append(movie_info_list)
    sht_name = re.sub("：", "-", title).split(" ") # Excel頁簽名稱不能包含：
    title_sht = sht_name[0].strip()
    wb.create_sheet(title_sht)
    ws = wb[title_sht]
    ws.cell(1, 1).value = title
    ws.cell(1, 1).font = font_header
    col = 0
    for dt in tr:
        dt_tr = dt.findAll("td")
        col += 1
        row = 2
        for rt in dt_tr:
            if not rt.a:
                ws.cell(row, col).value = rt.text
                ws.cell(row, col).font = xl.styles.Font(name='微軟正黑體', size=13, bold=True)
                ws.column_dimensions[ws.cell(row, col).column_letter].width = len(str(ws.cell(row, col).value))*2 + 4
                row += 1
            else:
                for a in rt:
                    ws.cell(row, col).value = a.text
                    ws.cell(row, col).font = font_other
                    row += 1
    wb["Sheet"]
    wb.save(path + filename)




def ad_font():
    wb = xl.load_workbook(path + filename)
    ws = wb.active
    col_all = ws["A:G"] # col_all是由A1~A6...G1~G6的4組tuple組成
    for ca in col_all:
        length = 2 * (max(len(str(ln.value)) for ln in ca)) + 4 # 列表推導式。從各組tuple中，取出字串個數最大的
        for fo in ca:
            fo.font = font_other
        if ca[0].column_letter in ["E", "F", "G"]:
            length = 4*(len(str(ca[0].value)))
        ws.column_dimensions[ca[0].column_letter].width = length

    col_header = ws["1"]
    for ch in col_header:
        ch.font = font_header
    wb.save(path + filename)


minshan_movie_info()




